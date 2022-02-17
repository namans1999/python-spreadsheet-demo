import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]
products_per_supplier = {}
total_value_per_supplier = {}
product_less_10 = {}

for product_row in range(2,product_list.max_row+1):
  supplier_name = product_list.cell(product_row,4).value
  inventory = product_list.cell(product_row,2).value
  price = product_list.cell(product_row,3).value
  product_no = product_list.cell(product_row,1).value
  inventory_price = product_list.cell(product_row,5)
  
  
  #calculation for no. of products
  if supplier_name in products_per_supplier:
    current_num_products = products_per_supplier.get(supplier_name)
    products_per_supplier[supplier_name] = current_num_products + 1
  else:
    print("adding a new supplier")
    products_per_supplier[supplier_name] = 1

  #calculate total inventory per supplier_name
  if supplier_name in total_value_per_supplier:
    current_total_value = total_value_per_supplier.get(supplier_name)
    total_value_per_supplier[supplier_name] = current_total_value + inventory * price
  else:
    total_value_per_supplier[supplier_name] = inventory * price

  #print all product with inventory less than 10
  if inventory < 10: 
    product_less_10[int(product_no)] = int(inventory)

  # adding new values to spreadsheet total inventory price
  inventory_price.value = inventory * price
  
    
inv_file.save("new file.xlsx")
print(product_less_10)

    
      